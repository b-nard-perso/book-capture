"""Export de la liste des livres vers un fichier xlsx.

Utilise openpyxl (open source, licence MIT).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Couleur d'en-tête (bleu clair)
COULEUR_ENTETE = "4472C4"


def exporter_xlsx(livres: list[dict], chemin_sortie: str | Path) -> Path:
    """Exporte la liste de livres dans un fichier xlsx.

    Le fichier contient trois colonnes : **Titre**, **Auteur**, **Quantité**.
    La première ligne est un en-tête formaté (fond coloré, police blanche en gras).
    Les colonnes sont auto-dimensionnées.

    Parameters
    ----------
    livres:
        Liste de dictionnaires avec les clés ``'titre'``, ``'auteur'`` et
        ``'quantite'``.
    chemin_sortie:
        Chemin du fichier xlsx à créer (sera écrasé s'il existe).

    Returns
    -------
    Path
        Chemin absolu du fichier créé.
    """
    chemin = Path(chemin_sortie)

    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = "Livres"

    # En-têtes de colonnes
    entetes = ["Titre", "Auteur", "Quantité"]
    style_entete_police = Font(bold=True, color="FFFFFF", size=12)
    style_entete_fond = PatternFill(fill_type="solid", fgColor=COULEUR_ENTETE)
    style_entete_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, texte in enumerate(entetes, start=1):
        cellule = feuille.cell(row=1, column=col_idx, value=texte)
        cellule.font = style_entete_police
        cellule.fill = style_entete_fond
        cellule.alignment = style_entete_align

    # Données
    style_donnees_align = Alignment(vertical="top", wrap_text=True)
    for ligne_idx, livre in enumerate(livres, start=2):
        feuille.cell(row=ligne_idx, column=1, value=livre.get("titre", "")).alignment = (
            style_donnees_align
        )
        feuille.cell(row=ligne_idx, column=2, value=livre.get("auteur", "")).alignment = (
            style_donnees_align
        )
        cellule_qte = feuille.cell(
            row=ligne_idx, column=3, value=livre.get("quantite", 1)
        )
        cellule_qte.alignment = Alignment(horizontal="center", vertical="top")

    # Ajustement automatique de la largeur des colonnes
    largeurs_min = [40, 30, 12]
    for col_idx, largeur_min in enumerate(largeurs_min, start=1):
        lettre = get_column_letter(col_idx)
        # Largeur basée sur le contenu + minimum garanti
        largeur_max = largeur_min
        for row in feuille.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    largeur_max = max(largeur_max, len(str(cell.value)) + 4)
        feuille.column_dimensions[lettre].width = min(largeur_max, 60)

    # Figer la ligne d'en-tête
    feuille.freeze_panes = "A2"

    classeur.save(chemin)
    return chemin.resolve()
