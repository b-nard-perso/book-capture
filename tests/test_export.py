"""Tests du module export."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from book_capture.export import exporter_xlsx


class TestExporterXlsx:
    """Tests de la fonction exporter_xlsx."""

    def test_fichier_cree(self, livres_consolides: list[dict], tmp_path: Path) -> None:
        """Le fichier xlsx doit être créé à l'emplacement demandé."""
        chemin = tmp_path / "livres.xlsx"
        exporter_xlsx(livres_consolides, chemin)
        assert chemin.exists()

    def test_retourne_chemin_absolu(self, livres_consolides: list[dict], tmp_path: Path) -> None:
        """La fonction doit retourner un chemin absolu."""
        chemin = tmp_path / "livres.xlsx"
        resultat = exporter_xlsx(livres_consolides, chemin)
        assert resultat.is_absolute()

    def test_feuille_nommee_livres(self, livres_consolides: list[dict], tmp_path: Path) -> None:
        """La feuille active du classeur doit s'appeler 'Livres'."""
        chemin = tmp_path / "livres.xlsx"
        exporter_xlsx(livres_consolides, chemin)
        classeur = openpyxl.load_workbook(chemin)
        assert "Livres" in classeur.sheetnames

    def test_entetes_corrects(self, livres_consolides: list[dict], tmp_path: Path) -> None:
        """La première ligne doit contenir les en-têtes Titre, Auteur, Quantité."""
        chemin = tmp_path / "livres.xlsx"
        exporter_xlsx(livres_consolides, chemin)
        classeur = openpyxl.load_workbook(chemin)
        feuille = classeur["Livres"]
        assert feuille.cell(1, 1).value == "Titre"
        assert feuille.cell(1, 2).value == "Auteur"
        assert feuille.cell(1, 3).value == "Quantité"

    def test_nombre_lignes_donnees(self, livres_consolides: list[dict], tmp_path: Path) -> None:
        """Le nombre de lignes de données doit correspondre au nombre de livres."""
        chemin = tmp_path / "livres.xlsx"
        exporter_xlsx(livres_consolides, chemin)
        classeur = openpyxl.load_workbook(chemin)
        feuille = classeur["Livres"]
        # Ligne 1 = en-tête, lignes 2..N = données
        lignes_donnees = feuille.max_row - 1
        assert lignes_donnees == len(livres_consolides)

    def test_quantite_correcte(self, livres_consolides: list[dict], tmp_path: Path) -> None:
        """Les quantités inscrites dans le fichier doivent correspondre aux données."""
        chemin = tmp_path / "livres.xlsx"
        exporter_xlsx(livres_consolides, chemin)
        classeur = openpyxl.load_workbook(chemin)
        feuille = classeur["Livres"]

        quantites_attendues = {l["titre"]: l["quantite"] for l in livres_consolides}
        for row in feuille.iter_rows(min_row=2, values_only=True):
            titre, auteur, quantite = row
            assert quantite == quantites_attendues[titre]

    def test_liste_vide_cree_fichier_avec_entetes(self, tmp_path: Path) -> None:
        """Une liste vide doit créer un fichier xlsx valide avec seulement les en-têtes."""
        chemin = tmp_path / "vide.xlsx"
        exporter_xlsx([], chemin)
        assert chemin.exists()
        classeur = openpyxl.load_workbook(chemin)
        feuille = classeur["Livres"]
        # Seulement la ligne d'en-tête
        assert feuille.max_row == 1

    def test_ecrase_fichier_existant(self, livres_consolides: list[dict], tmp_path: Path) -> None:
        """Si le fichier existe déjà, il doit être écrasé sans erreur."""
        chemin = tmp_path / "livres.xlsx"
        # Premier export
        exporter_xlsx([], chemin)
        # Deuxième export avec données
        exporter_xlsx(livres_consolides, chemin)
        classeur = openpyxl.load_workbook(chemin)
        feuille = classeur["Livres"]
        assert feuille.max_row == len(livres_consolides) + 1
